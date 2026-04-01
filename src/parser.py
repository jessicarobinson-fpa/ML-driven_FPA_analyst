"""
Parse the Close File HC tabs into normalized pandas DataFrames.

Reads by column-B label (not row number) so it handles the tab-specific
row offsets between HC/HC-US (summary at 166-169) and India/RD/Colleen
(summary at 170-173).
"""
from __future__ import annotations

import re
from datetime import datetime
from pathlib import Path

import openpyxl
import pandas as pd

from src.config import (
    CLOSE_FILE_PATH,
    DEPARTMENT_HIERARCHY,
    ENTITIES,
    HC_TABS,
    LC_FTE_ENTITIES,
    ONLY_ROW_SUFFIX,
    SUMMARY_ROW_LABELS,
    ZERO_HC_ROWS,
)

TAB_TO_ENTITY = {
    "HC": "Total",
    "HC-US": "US",
    "HC-India": "India",
    "HC-RD": "RD",
    "HC-Colleen": "Colleen",
}

_DEPT_TO_COST_CAT: dict[str, str] = {}
for _cat, _depts in DEPARTMENT_HIERARCHY["Entrata"].items():
    for _dept in _depts:
        _DEPT_TO_COST_CAT[_dept] = _cat


def _classify_level(label: str) -> str:
    """Return the hierarchy level for a row label."""
    if label == "Entrata":
        return "L0"
    if label in DEPARTMENT_HIERARCHY["Entrata"]:
        return "L1"
    if label in SUMMARY_ROW_LABELS:
        return "summary"
    if label.endswith(ONLY_ROW_SUFFIX):
        return "only"
    if " - " in label:
        return "L3"
    return "L2"


def _resolve_cost_category(label: str) -> str | None:
    """Map a row label to its L1 cost category."""
    base = label.split(" - ")[0].split(" (")[0].strip()
    if base in _DEPT_TO_COST_CAT:
        return _DEPT_TO_COST_CAT[base]
    if base in DEPARTMENT_HIERARCHY["Entrata"]:
        return base
    return None


def _parse_month(val) -> str | None:
    """Convert a month header like 'Mar-2026' to 'YYYY-MM' ISO string."""
    if val is None:
        return None
    s = str(val).strip()
    try:
        dt = datetime.strptime(s, "%b-%Y")
        return dt.strftime("%Y-%m")
    except ValueError:
        return None


def load_workbook(path: Path | None = None) -> openpyxl.Workbook:
    fp = path or CLOSE_FILE_PATH
    return openpyxl.load_workbook(fp, data_only=True)


def parse_actuals(wb: openpyxl.Workbook) -> pd.DataFrame:
    """
    Extract the trailing actuals grid (columns C-N) from all 5 HC tabs
    into a long-format DataFrame.

    Returns columns:
        tab, entity, department, cost_category, level, hc_type, month, value
    """
    rows: list[dict] = []

    for tab_name in HC_TABS:
        ws = wb[tab_name]
        entity = TAB_TO_ENTITY[tab_name]

        month_cols: list[tuple[int, str]] = []
        for col in range(3, 15):  # C=3 through N=14
            raw = ws.cell(row=4, column=col).value
            iso = _parse_month(raw)
            if iso:
                month_cols.append((col, iso))

        for row in range(6, ws.max_row + 1):
            label = ws.cell(row=row, column=2).value
            if label is None:
                continue
            label = str(label).strip()
            if not label:
                continue

            level = _classify_level(label)
            cost_cat = _resolve_cost_category(label)
            is_lc_fte = label == "Leasing Center - US"
            hc_type = "fte" if is_lc_fte else "integer"

            for col, month in month_cols:
                val = ws.cell(row=row, column=col).value
                rows.append({
                    "tab": tab_name,
                    "entity": entity,
                    "department": label,
                    "cost_category": cost_cat,
                    "level": level,
                    "hc_type": hc_type,
                    "month": month,
                    "value": val if val is not None else 0,
                })

    df = pd.DataFrame(rows)
    df["month"] = pd.to_datetime(df["month"])
    df["value"] = pd.to_numeric(df["value"], errors="coerce").fillna(0)
    return df


def parse_variance(wb: openpyxl.Workbook) -> pd.DataFrame:
    """
    Extract the month and EOY variance sections from all 5 HC tabs.

    Returns columns:
        tab, entity, department, cost_category, level, hc_type,
        variance_month,
        actuals, forecast, month_variance, month_commentary,
        cwm_fy, latest_forecast_fy, eoy_variance, eoy_commentary
    """
    rows: list[dict] = []

    for tab_name in HC_TABS:
        ws = wb[tab_name]
        entity = TAB_TO_ENTITY[tab_name]

        var_month_raw = ws.cell(row=4, column=16).value  # P4
        var_month = _parse_month(var_month_raw)

        for row in range(6, ws.max_row + 1):
            label = ws.cell(row=row, column=2).value
            if label is None:
                continue
            label = str(label).strip()
            if not label:
                continue

            level = _classify_level(label)
            cost_cat = _resolve_cost_category(label)
            is_lc_fte = label == "Leasing Center - US"
            hc_type = "fte" if is_lc_fte else "integer"

            actuals = ws.cell(row=row, column=16).value       # P
            forecast = ws.cell(row=row, column=17).value       # Q
            month_var = ws.cell(row=row, column=18).value      # R
            month_comm_raw = ws.cell(row=row, column=20).value  # T
            month_comm = month_comm_raw if isinstance(month_comm_raw, str) and month_comm_raw.strip() else None
            cwm_fy = ws.cell(row=row, column=22).value         # V
            lf_fy = ws.cell(row=row, column=23).value          # W
            eoy_var = ws.cell(row=row, column=24).value        # X
            eoy_comm_raw = ws.cell(row=row, column=26).value    # Z
            eoy_comm = eoy_comm_raw if isinstance(eoy_comm_raw, str) and eoy_comm_raw.strip() else None

            rows.append({
                "tab": tab_name,
                "entity": entity,
                "department": label,
                "cost_category": cost_cat,
                "level": level,
                "hc_type": hc_type,
                "variance_month": var_month,
                "actuals": actuals if actuals is not None else 0,
                "forecast": forecast if forecast is not None else 0,
                "month_variance": month_var if month_var is not None else 0,
                "month_commentary": month_comm,
                "cwm_fy": cwm_fy if cwm_fy is not None else 0,
                "latest_forecast_fy": lf_fy if lf_fy is not None else 0,
                "eoy_variance": eoy_var if eoy_var is not None else 0,
                "eoy_commentary": eoy_comm,
            })

    df = pd.DataFrame(rows)
    df["variance_month"] = pd.to_datetime(df["variance_month"])
    for col in ["actuals", "forecast", "month_variance",
                "cwm_fy", "latest_forecast_fy", "eoy_variance"]:
        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)
    return df


def filter_analytical(df: pd.DataFrame) -> pd.DataFrame:
    """
    Remove rows that carry no analytical signal:
    - (Only) placeholder rows
    - COR Only sub-rows (always zero)
    - Rows where level == 'only'
    """
    mask = (
        ~df["department"].str.endswith(ONLY_ROW_SUFFIX)
        & ~df["department"].isin(ZERO_HC_ROWS)
        & (df["level"] != "only")
    )
    return df[mask].reset_index(drop=True)
